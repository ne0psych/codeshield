"""
CodeShield Excel Report Generator
Produces .xlsx reports with dashboard summary, enriched findings,
SBOM, and remediation roadmap. All cells properly filled.
"""
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

logger = logging.getLogger("codeshield.reports.excel")

SEV_COLORS = {
    "CRITICAL": "DC143C", "HIGH": "FF6D00",
    "MEDIUM": "E6B400", "LOW": "2979FF", "INFO": "78909C",
}

# Consistent border style
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)

FINDING_HEADERS = [
    "ID", "Severity", "Plugin", "Title", "Description",
    "File Path", "Line", "Code Snippet", "Remediation",
    "Rule ID", "CVE ID", "CVSS", "CWE ID", "CWE Label",
    "OWASP Category", "OWASP Label", "MITRE ATT&CK ID", "MITRE Label",
    "PCI-DSS", "HIPAA", "Exploit Available", "False Positive",
    "Confidence", "Fix Effort (hrs)", "Trend",
    "Package", "Version", "Fixed Version", "License",
]

SBOM_HEADERS = ["Name", "Version", "PURL", "License", "Ecosystem"]


def generate_excel(scan: dict, findings: list, sbom: dict, project: dict = None) -> bytes:
    """Generate Excel report and return bytes."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "1F3864"
    _build_dashboard(ws, scan, findings, project=project)

    ws2 = wb.create_sheet("Findings")
    ws2.sheet_properties.tabColor = "DC143C"
    _build_findings(ws2, findings)

    if sbom and sbom.get("components"):
        ws3 = wb.create_sheet("SBOM")
        ws3.sheet_properties.tabColor = "00B050"
        _build_sbom(ws3, sbom)

    ws4 = wb.create_sheet("Remediation Roadmap")
    ws4.sheet_properties.tabColor = "7030A0"
    _build_roadmap(ws4, findings)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_fill():
    return PatternFill("solid", fgColor="1F3864")

def _header_font():
    return Font(bold=True, color="FFFFFF", size=10)


def _apply_headers(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def _build_dashboard(ws, scan, findings, project=None):
    """Build a clean dashboard summary sheet."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12

    # Title
    cell = ws.cell(row=1, column=1, value="CodeShield Security Report")
    cell.font = Font(bold=True, size=18, color="1F3864")
    ws.merge_cells("A1:B1")

    cell2 = ws.cell(row=2, column=1, value="Automated Security Scan Results")
    cell2.font = Font(size=10, color="666666", italic=True)
    ws.merge_cells("A2:B2")

    # Scan information section (with project context)
    ws.cell(row=4, column=1, value="SCAN INFORMATION").font = Font(bold=True, size=11, color="1F3864")
    ws.merge_cells("A4:B4")

    info_rows = []
    if project:
        info_rows.append(("Project", _safe(project.get("name", ""))))
        info_rows.append(("Project ID", _safe(project.get("project_id", ""))[:12]))
    info_rows.extend([
        ("Scan ID",     _safe(scan.get("scan_id", ""))),
        ("Filename",    _safe(scan.get("filename", ""))),
        ("Status",      _safe(scan.get("status", ""))),
        ("Scan Date",   _safe(scan.get("started_at", ""))),
        ("Completed",   _safe(scan.get("completed_at", ""))),
    ])
    for i, (label, value) in enumerate(info_rows, 5):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(bold=True, size=9)
        lc.fill = PatternFill("solid", fgColor="F0F0FA")
        lc.border = THIN_BORDER
        vc = ws.cell(row=i, column=2, value=value)
        vc.border = THIN_BORDER

    # Severity distribution section
    sev_row = 11
    ws.cell(row=sev_row, column=1, value="SEVERITY DISTRIBUTION").font = Font(
        bold=True, size=11, color="1F3864"
    )
    ws.merge_cells(f"A{sev_row}:B{sev_row}")

    # Headers for severity chart data
    sev_row += 1
    for col, h in enumerate(["Severity", "Count"], 1):
        c = ws.cell(row=sev_row, column=col + 3, value=h)
        c.font = _header_font()
        c.fill = _header_fill()
        c.border = THIN_BORDER

    sev_data = [
        ("Critical", scan.get("critical_count", 0), "DC143C"),
        ("High",     scan.get("high_count", 0),     "FF6D00"),
        ("Medium",   scan.get("medium_count", 0),   "E6B400"),
        ("Low",      scan.get("low_count", 0),      "2979FF"),
        ("Info",     scan.get("info_count", 0),      "78909C"),
    ]

    for i, (label, count, color) in enumerate(sev_data):
        r = sev_row + 1 + i
        # Visual severity badge in columns A-B
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, color="FFFFFF", size=10)
        lc.fill = PatternFill("solid", fgColor=color)
        lc.border = THIN_BORDER
        lc.alignment = Alignment(horizontal="center")

        vc = ws.cell(row=r, column=2, value=count)
        vc.font = Font(bold=True, size=14)
        vc.alignment = Alignment(horizontal="center")
        vc.border = THIN_BORDER

        # Chart data in columns D-E
        ws.cell(row=r, column=4, value=label).border = THIN_BORDER
        ws.cell(row=r, column=5, value=count).border = THIN_BORDER

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Findings by Severity"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    chart.style = 10

    data_ref = Reference(ws, min_col=5, min_row=sev_row,
                         max_row=sev_row + 5)
    cats_ref = Reference(ws, min_col=4, min_row=sev_row + 1,
                         max_row=sev_row + 5)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "D19")

    # Trends section
    trend_row = sev_row + 7
    ws.cell(row=trend_row, column=1, value="TREND ANALYSIS").font = Font(
        bold=True, size=11, color="1F3864"
    )
    ws.merge_cells(f"A{trend_row}:B{trend_row}")

    new_count = sum(1 for f in findings if f.get("trend_status") == "new")
    rec_count = sum(1 for f in findings if f.get("trend_status") == "recurring")
    total_effort = sum(_safe_float(f.get("fix_effort_hours", 0)) for f in findings)

    trend_data = [
        ("Total Findings", scan.get("total_findings", len(findings))),
        ("New Findings", new_count),
        ("Recurring Findings", rec_count),
        ("Est. Remediation Hours", round(total_effort, 1)),
    ]
    for i, (label, value) in enumerate(trend_data, trend_row + 1):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(bold=True, size=9)
        lc.fill = PatternFill("solid", fgColor="F0F0FA")
        lc.border = THIN_BORDER
        vc = ws.cell(row=i, column=2, value=value)
        vc.font = Font(bold=True, size=11)
        vc.border = THIN_BORDER
        vc.alignment = Alignment(horizontal="center")


def _build_findings(ws, findings):
    _apply_headers(ws, FINDING_HEADERS)

    for i, f in enumerate(findings, 2):
        row_data = [
            i - 1,
            _safe(f.get("severity", "")),
            _safe(f.get("plugin_name", "")),
            _safe(f.get("title", "")),
            _safe(f.get("description", ""))[:500],
            _safe(f.get("file_path", "")),
            f.get("line_number") or 0,
            _safe(f.get("code_snippet", ""))[:300],
            _safe(f.get("remediation", ""))[:500],
            _safe(f.get("rule_id", "")),
            _safe(f.get("cve_id", "")),
            f.get("cvss_score") or 0,
            _safe(f.get("cwe_id", "")),
            _safe(f.get("cwe_label", "")),
            _safe(f.get("owasp_category", "")),
            _safe(f.get("owasp_label", "")),
            _safe(f.get("mitre_id", "")),
            _safe(f.get("mitre_label", "")),
            _safe(f.get("pci_dss", "")),
            _safe(f.get("hipaa", "")),
            _safe(f.get("exploit_available", "")) or "unknown",
            "Yes" if f.get("false_positive") else "No",
            f.get("confidence") or 0.85,
            _safe_float(f.get("fix_effort_hours", 2.0)),
            _safe(f.get("trend_status", "")) or "new",
            _safe(f.get("package_name", "")),
            _safe(f.get("package_version", "")),
            _safe(f.get("fixed_version", "")),
            _safe(f.get("license_id", "")),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)

            if col == 2:  # severity color
                sev_str = str(val).upper()
                color = SEV_COLORS.get(sev_str, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, size=9,
                                 color="FFFFFF" if sev_str != "MEDIUM" else "000000")

    # Column widths
    widths = [5, 10, 12, 35, 30, 30, 6, 25, 30,
              12, 14, 6, 10, 20, 12, 20, 12, 20,
              8, 14, 12, 8, 8, 10, 8, 18, 10, 10, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_sbom(ws, sbom):
    _apply_headers(ws, SBOM_HEADERS)
    for i, c in enumerate(sbom.get("components", []), 2):
        vals = [
            _safe(c.get("name", "")),
            _safe(c.get("version", "")),
            _safe(c.get("purl", "")),
            _safe(c.get("license", "")),
            _safe(c.get("ecosystem", "")),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 30


def _build_roadmap(ws, findings):
    headers = ["Phase", "Severities", "Count", "Total Fix Hours",
               "Timeline", "Actions"]
    _apply_headers(ws, headers)
    phases = [
        ("Phase 1: Immediate", ["CRITICAL", "HIGH"], "This week",
         "Fix immediately. These represent active exploit risk."),
        ("Phase 2: Short-term", ["MEDIUM"], "This sprint",
         "Schedule within the current sprint or iteration."),
        ("Phase 3: Backlog", ["LOW", "INFO"], "Next quarter",
         "Add to technical debt backlog for future resolution."),
    ]
    for i, (label, sevs, timeline, action) in enumerate(phases, 2):
        matched = [f for f in findings if f.get("severity", "") in sevs]
        hours = sum(_safe_float(f.get("fix_effort_hours", 0)) for f in matched)

        vals = [label, ", ".join(sevs), len(matched),
                round(hours, 1), timeline, action]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)

    widths = [22, 18, 8, 14, 14, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _safe(val) -> str:
    """Convert value to safe string, handling None."""
    if val is None:
        return ""
    return str(val)


def _safe_float(val) -> float:
    """Convert value to float safely."""
    try:
        return float(val) if val else 0.0
    except (ValueError, TypeError):
        return 0.0
