#!/usr/bin/env python3
"""
CodeShield Report Generator
Produces a professional PDF report and Excel workbook from scan results.
"""

import os
import json
import datetime
from collections import defaultdict
from typing import List, Dict, Tuple
from dataclasses import asdict

# ─── PDF Report ───────────────────────────────────────────────────────────────

def generate_pdf_report(scan_results, sbom: dict, output_path: str, zip_name: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    SEV_COLORS = {
        "CRITICAL": colors.HexColor("#C0392B"),
        "HIGH":     colors.HexColor("#E67E22"),
        "MEDIUM":   colors.HexColor("#F1C40F"),
        "LOW":      colors.HexColor("#27AE60"),
        "INFO":     colors.HexColor("#3498DB"),
    }
    BG_DARK   = colors.HexColor("#1A1A2E")
    BG_MID    = colors.HexColor("#16213E")
    ACCENT    = colors.HexColor("#0F3460")
    TEXT_MAIN = colors.HexColor("#2C3E50")
    TEXT_LIGHT= colors.HexColor("#5D6D7E")
    WHITE     = colors.white

    styles = getSampleStyleSheet()
    base   = styles["Normal"]

    def S(name, **kw):
        return ParagraphStyle(name, parent=base, **kw)

    st_h1     = S("H1", fontSize=22, leading=28, textColor=BG_DARK,
                  fontName="Helvetica-Bold", spaceAfter=6)
    st_h2     = S("H2", fontSize=15, leading=20, textColor=ACCENT,
                  fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=10)
    st_h3     = S("H3", fontSize=11, leading=16, textColor=BG_MID,
                  fontName="Helvetica-Bold", spaceAfter=3, spaceBefore=6)
    st_body   = S("Body", fontSize=9, leading=13, textColor=TEXT_MAIN)
    st_small  = S("Small", fontSize=8, leading=12, textColor=TEXT_LIGHT)
    st_mono   = S("Mono", fontSize=8, leading=12, textColor=TEXT_MAIN,
                  fontName="Courier", backColor=colors.HexColor("#F8F9FA"))
    st_badge  = S("Badge", fontSize=8, leading=12, textColor=WHITE,
                  fontName="Helvetica-Bold", alignment=TA_CENTER)
    st_center = S("Center", fontSize=9, leading=13, alignment=TA_CENTER)
    st_title_white = S("TitleWhite", fontSize=28, leading=34, textColor=WHITE,
                       fontName="Helvetica-Bold")
    st_sub_white   = S("SubWhite", fontSize=11, leading=16, textColor=colors.HexColor("#BDC3C7"))

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title="CodeShield Security Report",
        author="CodeShield v1.0"
    )

    story = []
    all_vulns = [v for r in scan_results for v in r.vulnerabilities]
    sev_counts = defaultdict(int)
    for v in all_vulns:
        sev_counts[v.severity] += 1

    total = len(all_vulns)
    critical = sev_counts["CRITICAL"]
    high     = sev_counts["HIGH"]
    risk_score = min(100, critical*15 + high*5 +
                     sev_counts["MEDIUM"]*2 + sev_counts["LOW"])
    risk_label = ("CRITICAL RISK" if risk_score >= 60 else
                  "HIGH RISK" if risk_score >= 30 else
                  "MEDIUM RISK" if risk_score >= 10 else "LOW RISK")
    scan_date  = datetime.datetime.now().strftime("%d %B %Y  %H:%M UTC")

    # ── Cover Page ──
    cover_data = [[
        Paragraph("CodeShield", st_title_white),
        Paragraph("Security Assessment Report", st_sub_white),
        Spacer(1, 0.4*cm),
        Paragraph(f"Target: {zip_name}", st_sub_white),
        Paragraph(f"Scan Date: {scan_date}", st_sub_white),
        Spacer(1, 0.8*cm),
        Paragraph(f"Risk Score: {risk_score}/100  |  {risk_label}", ParagraphStyle(
            "RS", parent=base, fontSize=14, textColor=SEV_COLORS.get("CRITICAL" if risk_score>=60 else "HIGH" if risk_score>=30 else "MEDIUM", WHITE), fontName="Helvetica-Bold"
        )),
    ]]
    cover_table = Table([[
        [
            Paragraph("&#160;", st_sub_white), Spacer(1,1.5*cm),
            Paragraph("CodeShield", st_title_white),
            Paragraph("Security Assessment Report", st_sub_white),
            Spacer(1, 0.5*cm),
            Paragraph(f"Target: <b>{zip_name}</b>", st_sub_white),
            Paragraph(f"Scan Date: {scan_date}", st_sub_white),
            Spacer(1, 1*cm),
            Paragraph(f"Overall Risk Score: {risk_score}/100", ParagraphStyle(
                "RS2", parent=base, fontSize=18, textColor=WHITE, fontName="Helvetica-Bold")),
            Paragraph(risk_label, ParagraphStyle(
                "RL", parent=base, fontSize=13,
                textColor=SEV_COLORS.get("CRITICAL" if risk_score>=60 else "HIGH"),
                fontName="Helvetica-Bold")),
        ]
    ]], colWidths=[17*cm])
    cover_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), BG_DARK),
        ('ROWPADDING', (0,0), (-1,-1), 30),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 2, ACCENT),
    ]))
    story.append(cover_table)
    story.append(PageBreak())

    # ── Executive Summary ──
    story.append(Paragraph("Executive Summary", st_h1))
    story.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
    story.append(Spacer(1, 0.3*cm))

    summary_text = (
        f"CodeShield performed a comprehensive automated security assessment of <b>{zip_name}</b>. "
        f"The scan covered Static Application Security Testing (SAST), Software Composition Analysis (SCA), "
        f"Software Bill of Materials (SBOM), Infrastructure as Code (IaC), Container configuration, "
        f"and Secrets detection. "
        f"A total of <b>{total} vulnerabilities</b> were identified across {len(scan_results)} scan modules."
    )
    story.append(Paragraph(summary_text, st_body))
    story.append(Spacer(1, 0.5*cm))

    # Severity summary table
    sev_headers = ["Severity", "Count", "Risk Level", "Priority"]
    sev_rows = [
        ["CRITICAL", str(sev_counts["CRITICAL"]), "Immediate threat", "Fix within 24h"],
        ["HIGH",     str(sev_counts["HIGH"]),     "Significant risk", "Fix within 7 days"],
        ["MEDIUM",   str(sev_counts["MEDIUM"]),   "Moderate risk",    "Fix within 30 days"],
        ["LOW",      str(sev_counts["LOW"]),       "Minor risk",       "Fix in next release"],
        ["INFO",     str(sev_counts["INFO"]),      "Informational",    "Review as needed"],
        ["TOTAL",    str(total),                   "",                 ""],
    ]
    sev_table = Table(
        [sev_headers] + sev_rows,
        colWidths=[4*cm, 2.5*cm, 5*cm, 5.5*cm]
    )
    sev_style = [
        ('BACKGROUND', (0,0), (-1,0), BG_DARK),
        ('TEXTCOLOR',  (0,0), (-1,0), WHITE),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.HexColor("#FDFEFE"), colors.HexColor("#F2F3F4")]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor("#D5D8DC")),
        ('ALIGN',      (1,0), (1,-1), 'CENTER'),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#EBF5FB")),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0),(-1,-1), 5),
    ]
    for i, sev in enumerate(["CRITICAL","HIGH","MEDIUM","LOW","INFO"], 1):
        sev_style.append(('TEXTCOLOR', (0,i), (0,i), SEV_COLORS[sev]))
        sev_style.append(('FONTNAME',  (0,i), (0,i), 'Helvetica-Bold'))
    sev_table.setStyle(TableStyle(sev_style))
    story.append(sev_table)
    story.append(Spacer(1, 0.5*cm))

    # Scan module summary
    story.append(Paragraph("Scan Module Summary", st_h2))
    mod_headers = ["Module", "Tool", "Files Scanned", "Issues Found", "Duration"]
    mod_rows = [[r.scan_type, r.tool_used, str(r.files_scanned),
                 str(len(r.vulnerabilities)), f"{r.duration_sec}s"] for r in scan_results]
    mod_table = Table([mod_headers]+mod_rows, colWidths=[3*cm,5.5*cm,3*cm,3*cm,2.5*cm])
    mod_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), ACCENT),
        ('TEXTCOLOR',  (0,0), (-1,0), WHITE),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor("#FDFEFE"),colors.HexColor("#F2F3F4")]),
        ('GRID', (0,0),(-1,-1), 0.4, colors.HexColor("#D5D8DC")),
        ('TOPPADDING', (0,0),(-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]))
    story.append(mod_table)
    story.append(PageBreak())

    # ── Detailed Findings ──
    story.append(Paragraph("Detailed Findings", st_h1))
    story.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
    story.append(Spacer(1, 0.3*cm))

    # Group by scan type then severity
    by_type = defaultdict(list)
    for v in all_vulns:
        by_type[v.scan_type].append(v)

    sev_order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3,"INFO":4}
    finding_num = 0

    for scan_type, vulns in by_type.items():
        if not vulns:
            continue
        story.append(Paragraph(f"{scan_type} Findings", st_h2))
        sorted_vulns = sorted(vulns, key=lambda v: sev_order.get(v.severity, 5))

        for v in sorted_vulns:
            finding_num += 1
            sev_color = SEV_COLORS.get(v.severity, colors.grey)
            block = []

            # Header row
            header_data = [[
                Paragraph(f"Finding #{finding_num}: {v.title}", ParagraphStyle(
                    "FH", parent=base, fontSize=10, fontName="Helvetica-Bold",
                    textColor=WHITE)),
                Paragraph(v.severity, ParagraphStyle(
                    "SBadge", parent=base, fontSize=9, fontName="Helvetica-Bold",
                    textColor=WHITE, alignment=1)),
            ]]
            header_tbl = Table(header_data, colWidths=[13.5*cm, 2.5*cm])
            header_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,0), sev_color),
                ('BACKGROUND', (1,0), (1,0), sev_color),
                ('ROWPADDING', (0,0), (-1,-1), 6),
                ('VALIGN', (0,0),(-1,-1),'MIDDLE'),
            ]))
            block.append(header_tbl)

            # Details
            details = [
                ["Rule ID",       v.rule_id or "N/A"],
                ["Scan Type",     v.scan_type],
                ["File",          v.file_path or "N/A"],
                ["Line",          str(v.line_number) if v.line_number else "N/A"],
                ["CVE",           v.cve_id or "N/A"],
                ["CVSS Score",    str(v.cvss_score) if v.cvss_score else "N/A"],
            ]
            detail_tbl = Table(
                [[Paragraph(k, st_small), Paragraph(val, st_body)] for k,val in details],
                colWidths=[2.5*cm, 13.5*cm]
            )
            detail_tbl.setStyle(TableStyle([
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[colors.HexColor("#FDFEFE"),colors.HexColor("#F4F6F7")]),
                ('GRID',(0,0),(-1,-1),0.3,colors.HexColor("#D5D8DC")),
                ('TOPPADDING',(0,0),(-1,-1),3),
                ('BOTTOMPADDING',(0,0),(-1,-1),3),
                ('FONTSIZE',(0,0),(-1,-1),8),
            ]))
            block.append(detail_tbl)

            if v.description:
                block.append(Spacer(1,2*mm))
                block.append(Paragraph("<b>Description:</b>", st_small))
                block.append(Paragraph(v.description, st_body))

            if v.code_snippet:
                block.append(Spacer(1,2*mm))
                block.append(Paragraph("<b>Code Snippet:</b>", st_small))
                block.append(Paragraph(v.code_snippet.replace("<","&lt;").replace(">","&gt;"), st_mono))

            if v.remediation:
                block.append(Spacer(1,2*mm))
                block.append(Paragraph("<b>Remediation:</b>", st_small))
                block.append(Paragraph(v.remediation, ParagraphStyle(
                    "Fix", parent=base, fontSize=9, textColor=colors.HexColor("#1A5276"),
                    backColor=colors.HexColor("#EBF5FB"), leftIndent=8,
                    borderPadding=4)))

            block.append(Spacer(1, 4*mm))
            story.append(KeepTogether(block))

        story.append(PageBreak())

    # ── SBOM ──
    story.append(Paragraph("Software Bill of Materials (SBOM)", st_h1))
    story.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"CycloneDX SBOM — {len(sbom['components'])} components identified.", st_body))
    story.append(Spacer(1, 0.3*cm))

    if sbom["components"]:
        sbom_headers = ["#", "Package", "Version", "Type", "PURL"]
        sbom_rows = [[str(i+1), c["name"], c["version"], c["type"],
                      Paragraph(c["purl"][:55], st_small)]
                     for i, c in enumerate(sbom["components"][:60])]
        sbom_tbl = Table([sbom_headers]+sbom_rows,
                         colWidths=[0.8*cm, 4*cm, 2.5*cm, 2*cm, 7.7*cm])
        sbom_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0), ACCENT),
            ('TEXTCOLOR', (0,0),(-1,0), WHITE),
            ('FONTNAME',  (0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',  (0,0),(-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor("#FDFEFE"),colors.HexColor("#F2F3F4")]),
            ('GRID',(0,0),(-1,-1),0.3,colors.HexColor("#D5D8DC")),
            ('TOPPADDING',(0,0),(-1,-1),3),
            ('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        story.append(sbom_tbl)
        if len(sbom["components"]) > 60:
            story.append(Spacer(1,3*mm))
            story.append(Paragraph(
                f"... and {len(sbom['components'])-60} more components. See SBOM JSON for full list.",
                st_small))
    story.append(PageBreak())

    # ── Remediation Roadmap ──
    story.append(Paragraph("Remediation Roadmap", st_h1))
    story.append(HRFlowable(width="100%", thickness=2, color=ACCENT))
    story.append(Spacer(1, 0.3*cm))

    phases = [
        ("Phase 1 — Immediate (0-24 hours)", "CRITICAL",
         "Address all CRITICAL findings immediately. These represent active attack vectors."),
        ("Phase 2 — Short Term (1-7 days)", "HIGH",
         "Remediate all HIGH severity findings. These significantly elevate risk exposure."),
        ("Phase 3 — Medium Term (1-30 days)", "MEDIUM",
         "Address MEDIUM findings as part of the next sprint or release cycle."),
        ("Phase 4 — Low Priority (Next release)", "LOW",
         "Schedule LOW findings for resolution in upcoming maintenance windows."),
    ]
    for phase, sev, guidance in phases:
        phase_vulns = [v for v in all_vulns if v.severity == sev]
        story.append(Paragraph(phase, st_h2))
        story.append(Paragraph(
            f"{guidance} <b>({len(phase_vulns)} findings)</b>", st_body))
        if phase_vulns:
            for v in phase_vulns[:10]:
                story.append(Paragraph(
                    f"• [{v.scan_type}] {v.file_path}:{v.line_number} — {v.title}", st_small))
            if len(phase_vulns) > 10:
                story.append(Paragraph(f"  ... and {len(phase_vulns)-10} more.", st_small))
        story.append(Spacer(1, 4*mm))

    # Footer note
    story.append(PageBreak())
    story.append(Paragraph("Disclaimer", st_h2))
    story.append(Paragraph(
        "This report was generated by CodeShield v1.0 automated security scanning engine. "
        "Automated scanners may produce false positives. All findings should be reviewed by "
        "a qualified security professional before remediation. This report does not constitute "
        "a complete penetration test.",
        st_body))

    doc.build(story)
    print(f"  PDF report saved: {output_path}")


# ─── Excel Report ──────────────────────────────────────────────────────────────

def generate_excel_report(scan_results, sbom: dict, output_path: str, zip_name: str):
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint

    wb = openpyxl.Workbook()

    SEV_HEX = {
        "CRITICAL": "C0392B", "HIGH": "E67E22",
        "MEDIUM":   "F1C40F", "LOW":  "27AE60", "INFO": "3498DB"
    }
    HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1A1A2E")
    NORMAL_FONT  = Font(name="Calibri", size=9)
    BOLD_FONT    = Font(name="Calibri", bold=True, size=9)
    THIN_BORDER  = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC'))

    def style_header_row(ws, row, cols):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    all_vulns = [v for r in scan_results for v in r.vulnerabilities]
    from collections import defaultdict
    sev_counts = defaultdict(int)
    for v in all_vulns:
        sev_counts[v.severity] += 1

    # ── Sheet 1: Dashboard ──
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"CodeShield Security Report — {zip_name}"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1A1A2E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Generated: {datetime.datetime.now().strftime('%d %B %Y %H:%M')}  |  Total Vulnerabilities: {len(all_vulns)}"
    ws["A2"].font = Font(name="Calibri", size=10, color="5D6D7E")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["Severity", "Count", "Risk Level"])
    style_header_row(ws, 4, 3)
    sev_data = [
        ("CRITICAL", sev_counts["CRITICAL"], "Fix within 24h"),
        ("HIGH",     sev_counts["HIGH"],     "Fix within 7 days"),
        ("MEDIUM",   sev_counts["MEDIUM"],   "Fix within 30 days"),
        ("LOW",      sev_counts["LOW"],       "Next release"),
        ("INFO",     sev_counts["INFO"],      "Review"),
    ]
    for i, (sev, cnt, risk) in enumerate(sev_data, 5):
        ws.cell(row=i, column=1, value=sev).font = Font(name="Calibri", bold=True, size=9, color=SEV_HEX[sev])
        ws.cell(row=i, column=2, value=cnt).font = BOLD_FONT
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=risk).font = NORMAL_FONT
        for c in range(1,4):
            ws.cell(row=i, column=c).border = THIN_BORDER
            ws.cell(row=i, column=c).fill = PatternFill("solid", fgColor="FDFEFE" if i%2==0 else "F2F3F4")

    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Vulnerabilities by Severity"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    data_ref   = Reference(ws, min_col=2, min_row=4, max_row=9)
    cats_ref   = Reference(ws, min_col=1, min_row=5, max_row=9)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width = 15; chart.height = 10
    ws.add_chart(chart, "E4")

    # Module summary
    ws.cell(row=12, column=1, value="Scan Module Summary").font = TITLE_FONT
    ws.append(["Module", "Tool", "Files Scanned", "Issues", "Duration (s)"])
    style_header_row(ws, 13, 5)
    for r in scan_results:
        ws.append([r.scan_type, r.tool_used, r.files_scanned,
                   len(r.vulnerabilities), r.duration_sec])
    set_col_widths(ws, [12, 25, 14, 10, 12, 12, 12, 12])

    # ── Sheet 2: All Vulnerabilities ──
    ws2 = wb.create_sheet("All Vulnerabilities")
    ws2.sheet_view.showGridLines = False
    headers = ["#","Scan Type","Severity","Rule ID","Title","Description",
               "File","Line","Code Snippet","Remediation","CVE ID","CVSS"]
    ws2.append(headers)
    style_header_row(ws2, 1, len(headers))
    ws2.row_dimensions[1].height = 30

    sev_order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3,"INFO":4}
    sorted_vulns = sorted(all_vulns, key=lambda v: sev_order.get(v.severity,5))
    for i, v in enumerate(sorted_vulns, 1):
        row = [i, v.scan_type, v.severity, v.rule_id, v.title,
               v.description, v.file_path, v.line_number or "",
               v.code_snippet[:100], v.remediation, v.cve_id, v.cvss_score or ""]
        ws2.append(row)
        r = ws2.max_row
        sev_color = SEV_HEX.get(v.severity, "FFFFFF")
        ws2.cell(r, 3).fill = PatternFill("solid", fgColor=sev_color)
        ws2.cell(r, 3).font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        for c in range(1, len(headers)+1):
            ws2.cell(r, c).border = THIN_BORDER
            ws2.cell(r, c).font = NORMAL_FONT
            ws2.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        ws2.row_dimensions[r].height = 45

    set_col_widths(ws2, [4,10,10,10,25,35,25,6,30,40,12,6])

    # Auto filter
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_vulns)+1}"
    ws2.freeze_panes = "A2"

    # ── Sheets per scan type ──
    scan_sheets = {
        "SAST":"SAST Findings","SCA":"SCA Findings","IaC":"IaC Findings",
        "Container":"Container Findings","Secrets":"Secrets Findings"
    }
    for scan_type, sheet_name in scan_sheets.items():
        vulns = [v for v in all_vulns if v.scan_type == scan_type]
        if not vulns:
            continue
        ws_t = wb.create_sheet(sheet_name)
        ws_t.sheet_view.showGridLines = False
        ws_t.append(["#","Severity","Rule ID","Title","File","Line",
                     "Code Snippet","Remediation","CVE","CVSS"])
        style_header_row(ws_t, 1, 10)
        for i, v in enumerate(sorted(vulns, key=lambda v: sev_order.get(v.severity,5)), 1):
            ws_t.append([i, v.severity, v.rule_id, v.title, v.file_path,
                         v.line_number or "", v.code_snippet[:80],
                         v.remediation, v.cve_id, v.cvss_score or ""])
            r = ws_t.max_row
            sev_color = SEV_HEX.get(v.severity, "FFFFFF")
            ws_t.cell(r, 2).fill = PatternFill("solid", fgColor=sev_color)
            ws_t.cell(r, 2).font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
            for c in range(1, 11):
                ws_t.cell(r, c).border = THIN_BORDER
                ws_t.cell(r, c).font = NORMAL_FONT
                ws_t.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            ws_t.row_dimensions[r].height = 40
        set_col_widths(ws_t, [4,10,10,28,28,6,30,40,12,6])
        ws_t.auto_filter.ref = f"A1:J{len(vulns)+1}"
        ws_t.freeze_panes = "A2"

    # ── SBOM Sheet ──
    ws_sbom = wb.create_sheet("SBOM")
    ws_sbom.sheet_view.showGridLines = False
    ws_sbom.append(["#","Package Name","Version","Type","Supplier","PURL","Checksum"])
    style_header_row(ws_sbom, 1, 7)
    for i, c in enumerate(sbom["components"], 1):
        ws_sbom.append([i, c["name"], c["version"], c["type"],
                        c.get("supplier",""), c["purl"], c.get("checksum","")])
        r = ws_sbom.max_row
        fill = PatternFill("solid", fgColor="FDFEFE" if i%2==0 else "F2F3F4")
        for col in range(1, 8):
            ws_sbom.cell(r, col).border = THIN_BORDER
            ws_sbom.cell(r, col).font = NORMAL_FONT
            ws_sbom.cell(r, col).fill = fill
    set_col_widths(ws_sbom, [4,18,10,10,10,45,18])
    ws_sbom.auto_filter.ref = f"A1:G{len(sbom['components'])+1}"
    ws_sbom.freeze_panes = "A2"

    # ── Remediation Sheet ──
    ws_rem = wb.create_sheet("Remediation Plan")
    ws_rem.sheet_view.showGridLines = False
    ws_rem.merge_cells("A1:G1")
    ws_rem["A1"].value = "Remediation Roadmap"
    ws_rem["A1"].font = TITLE_FONT
    ws_rem["A1"].fill = PatternFill("solid", fgColor="0F3460")
    ws_rem["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws_rem["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_rem.row_dimensions[1].height = 30

    ws_rem.append(["Phase","Severity","Count","Timeline","Action"])
    style_header_row(ws_rem, 2, 5)
    phases = [
        (1,"CRITICAL",sev_counts["CRITICAL"],"0-24 hours","Patch/disable immediately. Escalate to security team."),
        (2,"HIGH",    sev_counts["HIGH"],     "1-7 days",  "Prioritise in current sprint. Review attack surface."),
        (3,"MEDIUM",  sev_counts["MEDIUM"],   "1-30 days", "Include in next release. Apply defence-in-depth."),
        (4,"LOW",     sev_counts["LOW"],       "Next cycle", "Schedule in backlog. Consider risk acceptance."),
    ]
    for phase, sev, cnt, timeline, action in phases:
        ws_rem.append([f"Phase {phase}", sev, cnt, timeline, action])
        r = ws_rem.max_row
        ws_rem.cell(r,2).fill = PatternFill("solid", fgColor=SEV_HEX[sev])
        ws_rem.cell(r,2).font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        for c in range(1,6):
            ws_rem.cell(r,c).border = THIN_BORDER
            ws_rem.cell(r,c).font = NORMAL_FONT
    set_col_widths(ws_rem, [10,12,8,14,55])

    wb.save(output_path)
    print(f"  Excel report saved: {output_path}")
